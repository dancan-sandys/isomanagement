from fastapi import APIRouter, Depends, HTTPException, Query, UploadFile, File, status
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from typing import Optional, List
from datetime import datetime
import csv

from app.core.database import get_db
from app.core.security import get_current_user
from app.core.permissions import require_permission_dependency, check_permission
from app.models.user import User
from app.models.rbac import Module, PermissionType
from app.models.audit_mgmt import (
    Audit as AuditModel, AuditStatus, AuditType,
    AuditChecklistTemplate, AuditChecklistItem, ChecklistResponse,
    AuditFinding, FindingSeverity, FindingStatus,
    AuditAttachment, AuditItemAttachment, AuditFindingAttachment, AuditAuditee,
	AuditPlan, AuditProgram, ProgramStatus, RiskMethod, AuditRisk, RiskLevel,
    AuditTeamMember, TeamMemberRole, CompetenceStatus,
    AuditEvidence, AuditActivityLog, AuditReportHistory,
)
from app.schemas.audit import (
    AuditCreate, AuditUpdate, AuditResponse, AuditListResponse,
    ChecklistTemplateCreate, ChecklistTemplateResponse,
    ChecklistItemCreate, ChecklistItemUpdate, ChecklistItemResponse,
    FindingCreate, FindingUpdate, FindingResponse,
    AuditItemAttachmentResponse, AuditFindingAttachmentResponse,
    AuditAttachmentResponse, AuditStatsResponse, AuditeeResponse,
	AuditKpisResponse, AuditPlanCreate, AuditPlanResponse,
    AuditProgramCreate, AuditProgramUpdate, AuditProgramResponse, AuditProgramListResponse, AuditProgramKpisResponse,
    AuditRiskCreate, AuditRiskUpdate, AuditRiskResponse, AuditRiskListResponse, RiskPlanResponse,
    AuditTeamMemberCreate, AuditTeamMemberUpdate, AuditTeamMemberResponse, AuditTeamMemberListResponse,
    AuditEvidenceCreate, AuditEvidenceUpdate, AuditEvidenceResponse, AuditEvidenceListResponse,
    AuditActivityLogCreate, AuditActivityLogUpdate, AuditActivityLogResponse, AuditActivityLogListResponse,
    CompetenceAssessmentRequest, CompetenceAssessmentResponse,
    ImpartialityCheckRequest, ImpartialityCheckResponse,
    FindingListResponse, FindingsAnalyticsResponse, BulkFindingsStatusUpdateRequest, BulkFindingsAssignRequest,
    AuditReportApproveRequest, AuditReportHistoryResponse,
)
from app.utils.audit import audit_event
from app.schemas.nonconformance import NonConformanceCreate, NonConformanceResponse
from app.models.nonconformance import NonConformance as NCModel, NonConformanceSource, NonConformanceStatus
from io import BytesIO
from openpyxl import Workbook
from reportlab.lib.pagesizes import A4
from reportlab.pdfgen import canvas
import logging

logger = logging.getLogger(__name__)
# Schedule conflict detection
@router.get("/schedule/conflicts", dependencies=[Depends(require_permission_dependency("audits:view"))])
async def detect_schedule_conflicts(
    start: Optional[datetime] = Query(None),
    end: Optional[datetime] = Query(None),
    auditor_id: Optional[int] = Query(None),
    department: Optional[str] = Query(None),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """
    Detect schedule conflicts based on overlapping audit date ranges for a given auditor or department.
    """
    q = db.query(AuditModel)
    if auditor_id:
        q = q.filter((AuditModel.auditor_id == auditor_id) | (AuditModel.lead_auditor_id == auditor_id))
    if department:
        q = q.filter(AuditModel.auditee_department == department)
    audits = q.all()
    conflicts = []
    # Simple O(n^2) overlap detection for current scale
    for i in range(len(audits)):
        a = audits[i]
        for j in range(i + 1, len(audits)):
            b = audits[j]
            a_start, a_end = getattr(a, 'start_date', None), getattr(a, 'end_date', None)
            b_start, b_end = getattr(b, 'start_date', None), getattr(b, 'end_date', None)
            if not a_start or not a_end or not b_start or not b_end:
                continue
            # Optional global window restriction
            if start and a_end < start and b_end < start:
                continue
            if end and a_start > end and b_start > end:
                continue
            # Overlap if ranges intersect
            if a_start <= b_end and b_start <= a_end:
                conflicts.append({
                    "audit_a": {"id": a.id, "title": a.title, "start": a_start, "end": a_end, "auditor_id": a.auditor_id, "lead_auditor_id": a.lead_auditor_id, "department": a.auditee_department},
                    "audit_b": {"id": b.id, "title": b.title, "start": b_start, "end": b_end, "auditor_id": b.auditor_id, "lead_auditor_id": b.lead_auditor_id, "department": b.auditee_department},
                    "reason": "overlap"
                })
    return {"total_conflicts": len(conflicts), "conflicts": conflicts}


@router.post("/schedule/bulk-update", dependencies=[Depends(require_permission_dependency("audits:update"))])
async def bulk_update_schedule(
    updates: List[dict],
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """
    Bulk update audit start/end dates. Each item: {id, start_date, end_date} (ISO strings).
    Enforces ownership: lead auditor or program manager.
    """
    updated = 0
    for u in updates:
        try:
            audit_id = int(u.get('id'))
        except Exception:
            continue
        audit = db.query(AuditModel).get(audit_id)
        if not audit:
            continue
        # Enforce stronger role for cross-team updates (MANAGE_PROGRAM)
        if not check_permission(current_user.id, Module.AUDITS, PermissionType.MANAGE_PROGRAM, db):
            # Allow lead auditor to move only their own audits
            if audit.lead_auditor_id != current_user.id:
                continue
        start_s = u.get('start_date')
        end_s = u.get('end_date')
        try:
            if start_s:
                audit.start_date = datetime.fromisoformat(start_s)
            if end_s:
                audit.end_date = datetime.fromisoformat(end_s)
            # Update governance fields
            audit.reschedule_count = (getattr(audit, 'reschedule_count', 0) or 0) + 1
            audit.last_rescheduled_at = datetime.utcnow()
            updated += 1
        except Exception:
            continue
    db.commit()
    return {"updated": updated}

router = APIRouter()


def can_perform_destructive_action(audit: AuditModel, current_user: User, db: Session) -> bool:
    """
    Check if user can perform destructive actions (delete, update) on an audit
    Returns True if user is lead auditor or program manager
    """
    # Super admin or audit manager can perform all actions
    if check_permission(current_user.id, Module.AUDITS, PermissionType.MANAGE_PROGRAM, db):
        return True
    
    # Lead auditor can perform destructive actions on their own audits
    if audit.lead_auditor_id == current_user.id:
        return True
    
    return False


def check_audit_ownership(audit: AuditModel, current_user: User, db: Session, action: str = "access") -> bool:
    """
    Check if user has ownership or appropriate permissions for audit actions
    Returns True if user can perform the action, raises HTTPException otherwise
    
    Action restrictions:
    - view/read: Lead auditor, team auditor, auditee, program manager
    - update: Lead auditor, program manager only
    - delete: Lead auditor, program manager only
    - create: Any user with audits:create permission
    """
    # Super admin or audit manager can perform all actions
    if check_permission(current_user.id, Module.AUDITS, PermissionType.MANAGE_PROGRAM, db):
        return True
    
    # For destructive operations (delete, update), only lead auditor or program manager
    if action in ["delete", "update"]:
        if audit.lead_auditor_id == current_user.id:
            return True
        else:
            logger.warning(
                f"Unauthorized destructive action attempt: User {current_user.id} ({current_user.username}) "
                f"attempted to {action} audit {audit.id} (Lead: {audit.lead_auditor_id})"
            )
            raise HTTPException(
                status_code=status.HTTP_403_FORBIDDEN,
                detail={
                    "error": "Access denied",
                    "message": f"Only the lead auditor or program manager can {action} this audit",
                    "audit_id": audit.id,
                    "audit_title": audit.title,
                    "required_role": "Lead Auditor or Program Manager",
                    "user_role": "User",
                    "action": action,
                    "lead_auditor_id": audit.lead_auditor_id
                }
            )
    
    # For view/read operations, allow lead auditor, team auditor, and auditee
    if action in ["view", "read"]:
        # Lead auditor can access their own audits
        if audit.lead_auditor_id == current_user.id:
            return True
        
        # Team auditor can access audits they're assigned to
        if audit.auditor_id == current_user.id:
            return True
        
        # Check if user is an auditee (read-only access)
        auditee = db.query(AuditAuditee).filter(
            AuditAuditee.audit_id == audit.id,
            AuditAuditee.user_id == current_user.id
        ).first()
        
        if auditee:
            return True
    
    # Log unauthorized access attempt
    logger.warning(
        f"Unauthorized audit access attempt: User {current_user.id} ({current_user.username}) "
        f"attempted to {action} audit {audit.id} (Lead: {audit.lead_auditor_id}, Team: {audit.auditor_id})"
    )
    
    raise HTTPException(
        status_code=status.HTTP_403_FORBIDDEN,
        detail={
            "error": "Access denied",
            "message": f"You do not have permission to {action} this audit",
            "audit_id": audit.id,
            "audit_title": audit.title,
            "required_role": "Lead Auditor, Team Auditor, or Auditee",
            "user_role": "User",
            "action": action
        }
    )


def check_finding_ownership(finding: AuditFinding, current_user: User, db: Session, action: str = "access") -> bool:
    """
    Check if user has ownership or appropriate permissions for finding actions
    
    Action restrictions:
    - view/read: Lead auditor, team auditor, auditee, program manager
    - update: Lead auditor, finding creator, program manager only
    - delete: Lead auditor, program manager only
    """
    # Get the audit to check ownership
    audit = db.query(AuditModel).filter(AuditModel.id == finding.audit_id).first()
    if not audit:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Audit not found")
    
    # Super admin or audit manager can perform all actions
    if check_permission(current_user.id, Module.AUDITS, PermissionType.MANAGE_PROGRAM, db):
        return True
    
    # For destructive operations (delete), only lead auditor or program manager
    if action == "delete":
        if audit.lead_auditor_id == current_user.id:
            return True
        else:
            logger.warning(
                f"Unauthorized finding deletion attempt: User {current_user.id} ({current_user.username}) "
                f"attempted to delete finding {finding.id} in audit {audit.id}"
            )
            raise HTTPException(
                status_code=status.HTTP_403_FORBIDDEN,
                detail={
                    "error": "Access denied",
                    "message": "Only the lead auditor or program manager can delete findings",
                    "finding_id": finding.id,
                    "audit_id": audit.id,
                    "required_role": "Lead Auditor or Program Manager",
                    "user_role": "User",
                    "action": action
                }
            )
    
    # For update operations, allow lead auditor, finding creator, or program manager
    if action == "update":
        if audit.lead_auditor_id == current_user.id:
            return True
        if finding.created_by == current_user.id:
            return True
        else:
            logger.warning(
                f"Unauthorized finding update attempt: User {current_user.id} ({current_user.username}) "
                f"attempted to update finding {finding.id} in audit {audit.id}"
            )
            raise HTTPException(
                status_code=status.HTTP_403_FORBIDDEN,
                detail={
                    "error": "Access denied",
                    "message": "Only the lead auditor, finding creator, or program manager can update findings",
                    "finding_id": finding.id,
                    "audit_id": audit.id,
                    "required_role": "Lead Auditor, Finding Creator, or Program Manager",
                    "user_role": "User",
                    "action": action,
                    "finding_creator_id": finding.created_by
                }
            )
    
    # For view/read operations, check audit ownership
    if action in ["view", "read"]:
        return check_audit_ownership(audit, current_user, db, action)
    
    return True


@router.get("/{audit_id:int}/permissions", dependencies=[Depends(require_permission_dependency("audits:view"))])
async def get_audit_permissions(audit_id: int, db: Session = Depends(get_db), current_user: User = Depends(get_current_user)):
    """
    Get user's permissions for a specific audit
    Useful for frontend to show/hide UI elements based on user permissions
    """
    audit = db.query(AuditModel).get(audit_id)
    if not audit:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Audit not found")
    
    # Check basic view access
    try:
        check_audit_ownership(audit, current_user, db, "view")
    except HTTPException:
        # If user can't even view, return minimal permissions
        return {
            "can_view": False,
            "can_update": False,
            "can_delete": False,
            "can_approve": False,
            "is_lead_auditor": False,
            "is_team_auditor": False,
            "is_auditee": False,
            "is_program_manager": False
        }
    
    # Check if user is program manager
    is_program_manager = check_permission(current_user.id, Module.AUDITS, PermissionType.MANAGE_PROGRAM, db)
    
    # Check if user is lead auditor
    is_lead_auditor = audit.lead_auditor_id == current_user.id
    
    # Check if user is team auditor
    is_team_auditor = audit.auditor_id == current_user.id
    
    # Check if user is auditee
    auditee = db.query(AuditAuditee).filter(
        AuditAuditee.audit_id == audit.id,
        AuditAuditee.user_id == current_user.id
    ).first()
    is_auditee = auditee is not None
    
    # Determine permissions based on role
    can_update = is_program_manager or is_lead_auditor
    can_delete = is_program_manager or is_lead_auditor
    can_approve = is_program_manager or is_lead_auditor
    
    return {
        "can_view": True,
        "can_update": can_update,
        "can_delete": can_delete,
        "can_approve": can_approve,
        "is_lead_auditor": is_lead_auditor,
        "is_team_auditor": is_team_auditor,
        "is_auditee": is_auditee,
        "is_program_manager": is_program_manager,
        "audit_id": audit_id,
        "lead_auditor_id": audit.lead_auditor_id,
        "team_auditor_id": audit.auditor_id
    }


@router.get("/", response_model=AuditListResponse, dependencies=[Depends(require_permission_dependency("audits:view"))])
async def list_audits(
    search: Optional[str] = Query(None),
    audit_type: Optional[str] = Query(None),
    status: Optional[str] = Query(None),
    department: Optional[str] = Query(None),
    auditor_id: Optional[int] = Query(None),
    page: int = Query(1, ge=1),
    size: int = Query(20, ge=1, le=100),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    q = db.query(AuditModel)
    if audit_type:
        q = q.filter(AuditModel.audit_type == AuditType(audit_type))
    if status:
        q = q.filter(AuditModel.status == AuditStatus(status))
    if search:
        like = f"%{search}%"
        q = q.filter(AuditModel.title.ilike(like))
    if department:
        q = q.filter(AuditModel.auditee_department == department)
    if auditor_id:
        q = q.filter((AuditModel.auditor_id == auditor_id) | (AuditModel.lead_auditor_id == auditor_id))
    total = q.count()
    items = q.order_by(AuditModel.created_at.desc()).offset((page - 1) * size).limit(size).all()
    return AuditListResponse(items=items, total=total, page=page, size=size, pages=(total + size - 1) // size)


@router.post("/", response_model=AuditResponse, dependencies=[Depends(require_permission_dependency("audits:create"))])
async def create_audit(
    payload: AuditCreate,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    audit = AuditModel(
        title=payload.title,
        audit_type=AuditType(payload.audit_type),
        scope=payload.scope,
        objectives=payload.objectives,
        criteria=payload.criteria,
        start_date=payload.start_date,
        end_date=payload.end_date,
        status=AuditStatus(payload.status or "planned"),
        auditor_id=payload.auditor_id,
        lead_auditor_id=payload.lead_auditor_id,
        auditee_department=payload.auditee_department,
        created_by=current_user.id,
    )
    db.add(audit)
    db.commit()
    db.refresh(audit)
    try:
        audit_event(db, current_user.id, "audit_created", "audits", str(audit.id))
    except Exception:
        pass
    return audit


@router.get("/{audit_id:int}", response_model=AuditResponse, dependencies=[Depends(require_permission_dependency("audits:view"))])
async def get_audit(audit_id: int, db: Session = Depends(get_db), current_user: User = Depends(get_current_user)):
    audit = db.query(AuditModel).get(audit_id)
    if not audit:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Audit not found")
    
    # Check ownership for view access
    check_audit_ownership(audit, current_user, db, "view")
    
    return audit


@router.put("/{audit_id:int}", response_model=AuditResponse, dependencies=[Depends(require_permission_dependency("audits:update"))])
async def update_audit(audit_id: int, payload: AuditUpdate, db: Session = Depends(get_db), current_user: User = Depends(get_current_user)):
    audit = db.query(AuditModel).get(audit_id)
    if not audit:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Audit not found")
    
    # Check ownership for update access
    check_audit_ownership(audit, current_user, db, "update")
    
    for field, value in payload.model_dump(exclude_unset=True).items():
        if field in {"audit_type", "status"} and value is not None:
            value = (AuditType(value) if field == "audit_type" else AuditStatus(value))
        setattr(audit, field, value)
    db.commit()
    db.refresh(audit)
    try:
        audit_event(db, current_user.id, "audit_updated", "audits", str(audit.id))
    except Exception:
        pass
    return audit


@router.delete("/{audit_id:int}", dependencies=[Depends(require_permission_dependency("audits:delete"))])
async def delete_audit(audit_id: int, db: Session = Depends(get_db), current_user: User = Depends(get_current_user)):
    audit = db.query(AuditModel).get(audit_id)
    if not audit:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Audit not found")
    
    # Check ownership for delete access
    check_audit_ownership(audit, current_user, db, "delete")
    
    db.delete(audit)
    db.commit()
    try:
        audit_event(db, current_user.id, "audit_deleted", "audits", str(audit_id))
    except Exception:
        pass
    return {"message": "Audit deleted"}


# Checklist template endpoints
@router.post("/templates", response_model=ChecklistTemplateResponse)
async def create_template(payload: ChecklistTemplateCreate, db: Session = Depends(get_db), current_user: User = Depends(get_current_user)):
    template = AuditChecklistTemplate(
        name=payload.name,
        clause_ref=payload.clause_ref,
        question=payload.question,
        requirement=payload.requirement,
        category=payload.category,
        is_active=payload.is_active if payload.is_active is not None else True,
        created_by=current_user.id,
    )
    db.add(template)
    db.commit()
    db.refresh(template)
    return template


@router.get("/templates", response_model=List[ChecklistTemplateResponse])
async def list_templates(
    include_inactive: bool = Query(False),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    q = db.query(AuditChecklistTemplate)
    if not include_inactive:
        q = q.filter(AuditChecklistTemplate.is_active == True)
    return q.order_by(AuditChecklistTemplate.created_at.desc()).all()


@router.post("/templates/{template_id}/activate", response_model=ChecklistTemplateResponse)
async def activate_template(template_id: int, db: Session = Depends(get_db), current_user: User = Depends(get_current_user)):
    t = db.query(AuditChecklistTemplate).get(template_id)
    if not t:
        raise HTTPException(status_code=404, detail="Template not found")
    t.is_active = True
    db.commit()
    db.refresh(t)
    try:
        audit_event(db, current_user.id, "audit_template_activated", "audits", str(template_id))
    except Exception:
        pass
    return t


@router.post("/templates/{template_id}/deactivate", response_model=ChecklistTemplateResponse)
async def deactivate_template(template_id: int, db: Session = Depends(get_db), current_user: User = Depends(get_current_user)):
    t = db.query(AuditChecklistTemplate).get(template_id)
    if not t:
        raise HTTPException(status_code=404, detail="Template not found")
    t.is_active = False
    db.commit()
    db.refresh(t)
    try:
        audit_event(db, current_user.id, "audit_template_deactivated", "audits", str(template_id))
    except Exception:
        pass
    return t


@router.post("/templates/import", response_model=dict)
async def import_templates_csv(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    # Expect CSV headers: name, clause_ref, question, requirement, category, is_active(optional)
    content = (await file.read()).decode("utf-8-sig", errors="ignore")
    reader = csv.DictReader(content.splitlines())
    created = 0
    updated = 0
    skipped = 0
    for row in reader:
        name = (row.get("name") or "").strip()
        question = (row.get("question") or "").strip()
        if not name or not question:
            skipped += 1
            continue
        clause_ref = (row.get("clause_ref") or "").strip() or None
        requirement = (row.get("requirement") or "").strip() or None
        category = (row.get("category") or "").strip() or None
        is_active_val = (row.get("is_active") or "").strip().lower()
        is_active = True if is_active_val in {"", "1", "true", "yes", "y"} else False

        # Upsert by (name, clause_ref, question)
        existing = (
            db.query(AuditChecklistTemplate)
            .filter(AuditChecklistTemplate.name == name)
            .filter(AuditChecklistTemplate.clause_ref == clause_ref)
            .filter(AuditChecklistTemplate.question == question)
            .first()
        )
        if existing:
            existing.requirement = requirement
            existing.category = category
            existing.is_active = is_active
            updated += 1
        else:
            t = AuditChecklistTemplate(
                name=name,
                clause_ref=clause_ref,
                question=question,
                requirement=requirement,
                category=category,
                is_active=is_active,
                created_by=current_user.id,
            )
            db.add(t)
            created += 1
    db.commit()
    try:
        audit_event(db, current_user.id, "audit_templates_imported", "audits", "templates", {"created": created, "updated": updated, "skipped": skipped})
    except Exception:
        pass
    return {"created": created, "updated": updated, "skipped": skipped}


# Checklist items
@router.get("/{audit_id:int}/checklist", response_model=List[ChecklistItemResponse], dependencies=[Depends(require_permission_dependency("audits:view"))])
async def list_checklist_items(audit_id: int, db: Session = Depends(get_db), current_user: User = Depends(get_current_user)):
    audit = db.query(AuditModel).get(audit_id)
    if not audit:
        raise HTTPException(status_code=404, detail="Audit not found")
    items = db.query(AuditChecklistItem).filter(AuditChecklistItem.audit_id == audit_id).order_by(AuditChecklistItem.created_at.asc()).all()
    return items
@router.post("/{audit_id:int}/checklist", response_model=ChecklistItemResponse, dependencies=[Depends(require_permission_dependency("audits:create"))])
async def add_checklist_item(audit_id: int, payload: ChecklistItemCreate, db: Session = Depends(get_db), current_user: User = Depends(get_current_user)):
    audit = db.query(AuditModel).get(audit_id)
    if not audit:
        raise HTTPException(status_code=404, detail="Audit not found")
    item = AuditChecklistItem(
        audit_id=audit_id,
        template_id=payload.template_id,
        clause_ref=payload.clause_ref,
        question=payload.question,
        response=ChecklistResponse(payload.response) if payload.response else None,
        evidence_text=payload.evidence_text,
        responsible_person_id=payload.responsible_person_id,
        due_date=payload.due_date,
    )
    db.add(item)
    db.commit()
    db.refresh(item)
    return item


@router.put("/checklist/{item_id}", response_model=ChecklistItemResponse, dependencies=[Depends(require_permission_dependency("audits:update"))])
async def update_checklist_item(item_id: int, payload: ChecklistItemUpdate, db: Session = Depends(get_db), current_user: User = Depends(get_current_user)):
    item = db.query(AuditChecklistItem).get(item_id)
    if not item:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Checklist item not found")
    
    # Get the audit to check ownership
    audit = db.query(AuditModel).get(item.audit_id)
    if not audit:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Audit not found")
    
    # Check ownership for checklist item update
    check_audit_ownership(audit, current_user, db, "update")
    
    for field, value in payload.model_dump(exclude_unset=True).items():
        if field == "response" and value is not None:
            value = ChecklistResponse(value)
        setattr(item, field, value)
    db.commit()
    db.refresh(item)
    return item


# Findings
@router.get("/{audit_id:int}/findings", response_model=List[FindingResponse], dependencies=[Depends(require_permission_dependency("audits:view"))])
async def list_findings(audit_id: int, db: Session = Depends(get_db), current_user: User = Depends(get_current_user)):
    audit = db.query(AuditModel).get(audit_id)
    if not audit:
        raise HTTPException(status_code=404, detail="Audit not found")
    f = db.query(AuditFinding).filter(AuditFinding.audit_id == audit_id).order_by(AuditFinding.created_at.asc()).all()
    return f
@router.post("/{audit_id:int}/findings", response_model=FindingResponse, dependencies=[Depends(require_permission_dependency("audits:create"))])
async def add_finding(audit_id: int, payload: FindingCreate, db: Session = Depends(get_db), current_user: User = Depends(get_current_user)):
    audit = db.query(AuditModel).get(audit_id)
    if not audit:
        raise HTTPException(status_code=404, detail="Audit not found")
    finding = AuditFinding(
        audit_id=audit_id,
        clause_ref=payload.clause_ref,
        description=payload.description,
        severity=FindingSeverity(payload.severity),
        corrective_action=payload.corrective_action,
        responsible_person_id=payload.responsible_person_id,
        target_completion_date=payload.target_completion_date,
        status=FindingStatus(payload.status or "open"),
        related_nc_id=payload.related_nc_id,
    )
    db.add(finding)
    db.commit()
    db.refresh(finding)
    return finding


@router.put("/findings/{finding_id}", response_model=FindingResponse, dependencies=[Depends(require_permission_dependency("audits:update"))])
async def update_finding(finding_id: int, payload: FindingUpdate, db: Session = Depends(get_db), current_user: User = Depends(get_current_user)):
    finding = db.query(AuditFinding).get(finding_id)
    if not finding:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Finding not found")
    
    # Check ownership for finding update
    check_finding_ownership(finding, current_user, db, "update")
    
    updated_fields = payload.model_dump(exclude_unset=True)
    for field, value in updated_fields.items():
        if field == "severity" and value is not None:
            value = FindingSeverity(value)
        if field == "status" and value is not None:
            value = FindingStatus(value)
        setattr(finding, field, value)
    # Handle closed_at stamping
    try:
        if 'status' in updated_fields:
            if finding.status in [FindingStatus.VERIFIED, FindingStatus.CLOSED] and not finding.closed_at:
                finding.closed_at = datetime.utcnow()
            if finding.status in [FindingStatus.OPEN, FindingStatus.IN_PROGRESS]:
                finding.closed_at = None
    except Exception:
        pass
    db.commit()
    db.refresh(finding)
    return finding


# Cross-audit findings aggregation
@router.get("/findings", response_model=FindingListResponse, dependencies=[Depends(require_permission_dependency("audits:view"))])
async def list_all_findings(
    audit_id: Optional[int] = Query(None),
    program_id: Optional[int] = Query(None),
    severity: Optional[str] = Query(None),
    status: Optional[str] = Query(None),
    department: Optional[str] = Query(None),
    responsible_person_id: Optional[int] = Query(None),
    has_nc: Optional[bool] = Query(None),
    overdue: Optional[bool] = Query(None),
    created_from: Optional[datetime] = Query(None),
    created_to: Optional[datetime] = Query(None),
    page: int = Query(1, ge=1),
    size: int = Query(20, ge=1, le=200),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    q = db.query(AuditFinding)
    if audit_id:
        q = q.filter(AuditFinding.audit_id == audit_id)
    if program_id:
        # join via AuditModel
        q = q.join(AuditModel, AuditModel.id == AuditFinding.audit_id)
        q = q.filter(AuditModel.program_id == program_id)
    if severity:
        try:
            q = q.filter(AuditFinding.severity == FindingSeverity(severity))
        except Exception:
            pass
    if status:
        try:
            q = q.filter(AuditFinding.status == FindingStatus(status))
        except Exception:
            pass
    if department:
        q = q.join(AuditModel, AuditModel.id == AuditFinding.audit_id) if 'AuditModel' not in str(q) else q
        q = q.filter(AuditModel.auditee_department == department)
    if responsible_person_id:
        q = q.filter(AuditFinding.responsible_person_id == responsible_person_id)
    if has_nc is not None:
        if has_nc:
            q = q.filter(AuditFinding.related_nc_id.isnot(None))
        else:
            q = q.filter(AuditFinding.related_nc_id.is_(None))
    if created_from:
        q = q.filter(AuditFinding.created_at >= created_from)
    if created_to:
        q = q.filter(AuditFinding.created_at <= created_to)
    if overdue is not None:
        now = datetime.utcnow()
        if overdue:
            q = q.filter(
                (AuditFinding.status.notin_([FindingStatus.VERIFIED, FindingStatus.CLOSED])) &
                (AuditFinding.target_completion_date.isnot(None)) &
                (AuditFinding.target_completion_date < now)
            )
        else:
            q = q.filter(
                (AuditFinding.status.in_([FindingStatus.VERIFIED, FindingStatus.CLOSED])) |
                (AuditFinding.target_completion_date.is_(None)) |
                (AuditFinding.target_completion_date >= now)
            )

    total = q.count()
    items = (
        q.order_by(AuditFinding.created_at.desc())
         .offset((page - 1) * size)
         .limit(size)
         .all()
    )
    return FindingListResponse(items=items, total=total, page=page, size=size, pages=(total + size - 1) // size)


@router.post("/findings/bulk-update-status", dependencies=[Depends(require_permission_dependency("audits:update"))])
async def bulk_update_findings_status(
    payload: BulkFindingsStatusUpdateRequest,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    # Update with ownership check per finding's audit
    updated = 0
    for fid in payload.finding_ids:
        finding = db.query(AuditFinding).get(fid)
        if not finding:
            continue
        audit = db.query(AuditModel).get(finding.audit_id)
        if not audit:
            continue
        try:
            check_finding_ownership(finding, current_user, db, "update")
        except HTTPException:
            continue
        finding.status = FindingStatus(payload.status)
        if finding.status in [FindingStatus.VERIFIED, FindingStatus.CLOSED] and not finding.closed_at:
            finding.closed_at = datetime.utcnow()
        updated += 1
    db.commit()
    return {"updated": updated}


@router.post("/findings/bulk-assign", dependencies=[Depends(require_permission_dependency("audits:update"))])
async def bulk_assign_findings(
    payload: BulkFindingsAssignRequest,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    assigned = 0
    for fid in payload.finding_ids:
        finding = db.query(AuditFinding).get(fid)
        if not finding:
            continue
        audit = db.query(AuditModel).get(finding.audit_id)
        if not audit:
            continue
        try:
            check_finding_ownership(finding, current_user, db, "update")
        except HTTPException:
            continue
        finding.responsible_person_id = payload.responsible_person_id
        assigned += 1
    db.commit()
    return {"assigned": assigned}


@router.get("/findings/analytics", response_model=FindingsAnalyticsResponse, dependencies=[Depends(require_permission_dependency("audits:view"))])
async def findings_analytics(
    program_id: Optional[int] = Query(None),
    department: Optional[str] = Query(None),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    q = db.query(AuditFinding)
    if program_id:
        q = q.join(AuditModel, AuditModel.id == AuditFinding.audit_id).filter(AuditModel.program_id == program_id)
    if department:
        q = q.join(AuditModel, AuditModel.id == AuditFinding.audit_id).filter(AuditModel.auditee_department == department)
    findings = q.all()
    by_severity = {s.value: 0 for s in FindingSeverity}
    by_status = {s.value: 0 for s in FindingStatus}
    for f in findings:
        try:
            by_severity[f.severity.value if hasattr(f.severity, 'value') else str(f.severity)] += 1
        except Exception:
            pass
        try:
            by_status[f.status.value if hasattr(f.status, 'value') else str(f.status)] += 1
        except Exception:
            pass
    now = datetime.utcnow()
    open_findings = len([f for f in findings if f.status in [FindingStatus.OPEN, FindingStatus.IN_PROGRESS]])
    overdue_findings = len([f for f in findings if f.status not in [FindingStatus.VERIFIED, FindingStatus.CLOSED] and f.target_completion_date and f.target_completion_date < now])
    critical_findings = len([f for f in findings if f.severity == FindingSeverity.CRITICAL])
    durations = []
    for f in findings:
        if f.closed_at and f.created_at:
            try:
                durations.append((f.closed_at - f.created_at).days)
            except Exception:
                pass
    average_closure_days = (sum(durations) / len(durations)) if durations else None
    return FindingsAnalyticsResponse(
        by_severity=by_severity,
        by_status=by_status,
        open_findings=open_findings,
        overdue_findings=overdue_findings,
        critical_findings=critical_findings,
        average_closure_days=average_closure_days,
    )


# Attachments
@router.get("/{audit_id:int}/attachments", response_model=List[AuditAttachmentResponse], dependencies=[Depends(require_permission_dependency("audits:view"))])
async def list_attachments(audit_id: int, db: Session = Depends(get_db), current_user: User = Depends(get_current_user)):
    audit = db.query(AuditModel).get(audit_id)
    if not audit:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Audit not found")
    
    # Check ownership for view access
    check_audit_ownership(audit, current_user, db, "view")
    
    return (
        db.query(AuditAttachment)
        .filter(AuditAttachment.audit_id == audit_id)
        .order_by(AuditAttachment.uploaded_at.desc())
        .all()
    )

@router.post("/{audit_id:int}/attachments", response_model=dict, dependencies=[Depends(require_permission_dependency("audits:create"))])
async def upload_attachment(audit_id: int, file: UploadFile = File(...), db: Session = Depends(get_db), current_user: User = Depends(get_current_user)):
    audit = db.query(AuditModel).get(audit_id)
    if not audit:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Audit not found")
    
    # Check ownership for upload access
    check_audit_ownership(audit, current_user, db, "update")
    
    # Use storage service for secure file handling
    storage_service = StorageService()
    try:
        file_path, file_size, content_type, original_filename, checksum = storage_service.save_upload(
            file, subdir="audits"
        )
    except HTTPException as e:
        raise e
    except Exception as e:
        logger.error(f"File upload failed for audit {audit_id}: {str(e)}")
        raise HTTPException(status_code=status.HTTP_500_INTERNAL_SERVER_ERROR, detail="File upload failed")
    
    attachment = AuditAttachment(
        audit_id=audit_id,
        file_path=file_path,
        filename=original_filename,
        uploaded_by=current_user.id,
    )
    db.add(attachment)
    db.commit()
    try:
        audit_event(db, current_user.id, "audit_attachment_uploaded", "audits", str(audit_id))
    except Exception:
        pass
    return {"file_path": file_path, "filename": original_filename, "size": file_size, "checksum": checksum}


@router.get("/attachments/{attachment_id}/download")
async def download_attachment(attachment_id: int, db: Session = Depends(get_db), current_user: User = Depends(get_current_user)):
    att = db.query(AuditAttachment).get(attachment_id)
    if not att:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Attachment not found")
    
    # Get the audit to check ownership
    audit = db.query(AuditModel).get(att.audit_id)
    if not audit:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Audit not found")
    
    # Check ownership for download access
    check_audit_ownership(audit, current_user, db, "view")
    
    # Use storage service for secure file download
    storage_service = StorageService()
    return storage_service.create_file_response(att.file_path, att.filename)


@router.delete("/attachments/{attachment_id}", dependencies=[Depends(require_permission_dependency("audits:delete"))])
async def delete_attachment(attachment_id: int, db: Session = Depends(get_db), current_user: User = Depends(get_current_user)):
    att = db.query(AuditAttachment).get(attachment_id)
    if not att:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Attachment not found")
    
    # Get the audit to check ownership
    audit = db.query(AuditModel).get(att.audit_id)
    if not audit:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Audit not found")
    
    # Check ownership for delete access
    check_audit_ownership(audit, current_user, db, "delete")
    
    # Use storage service to delete the file
    storage_service = StorageService()
    storage_service.delete_file(att.file_path)
    
    db.delete(att)
    db.commit()
    return {"message": "Attachment deleted"}


# Upload attachment for a checklist item
@router.post("/checklist/{item_id}/attachments", response_model=AuditItemAttachmentResponse, dependencies=[Depends(require_permission_dependency("audits:create"))])
async def upload_item_attachment(
    item_id: int,
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    item = db.query(AuditChecklistItem).get(item_id)
    if not item:
        raise HTTPException(status_code=404, detail="Checklist item not found")
    
    # Use storage service for secure file handling
    storage_service = StorageService()
    try:
        file_path, file_size, content_type, original_filename, checksum = storage_service.save_upload(
            file, subdir="audits/items"
        )
    except HTTPException as e:
        raise e
    except Exception as e:
        raise HTTPException(status_code=500, detail="File upload failed")
    
    attachment = AuditItemAttachment(
        item_id=item_id,
        file_path=file_path,
        filename=original_filename,
        uploaded_by=current_user.id,
    )
    db.add(attachment)
    db.commit()
    db.refresh(attachment)
    try:
        audit_event(db, current_user.id, "audit_item_attachment_uploaded", "audits", str(item.audit_id))
    except Exception:
        pass
    return attachment


# Upload attachment for a finding
@router.post("/findings/{finding_id}/attachments", response_model=AuditFindingAttachmentResponse, dependencies=[Depends(require_permission_dependency("audits:create"))])
async def upload_finding_attachment(
    finding_id: int,
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    finding = db.query(AuditFinding).get(finding_id)
    if not finding:
        raise HTTPException(status_code=404, detail="Finding not found")
    
    # Use storage service for secure file handling
    storage_service = StorageService()
    try:
        file_path, file_size, content_type, original_filename, checksum = storage_service.save_upload(
            file, subdir="audits/findings"
        )
    except HTTPException as e:
        raise e
    except Exception as e:
        raise HTTPException(status_code=500, detail="File upload failed")
    
    attachment = AuditFindingAttachment(
        finding_id=finding_id,
        file_path=file_path,
        filename=original_filename,
        uploaded_by=current_user.id,
    )
    db.add(attachment)
    db.commit()
    db.refresh(attachment)
    try:
        audit_event(db, current_user.id, "audit_finding_attachment_uploaded", "audits", str(finding.audit_id))
    except Exception:
        pass
    return attachment


# List/download/delete attachments for checklist item
@router.get("/checklist/{item_id}/attachments", response_model=List[AuditItemAttachmentResponse])
async def list_item_attachments(item_id: int, db: Session = Depends(get_db), current_user: User = Depends(get_current_user)):
    item = db.query(AuditChecklistItem).get(item_id)
    if not item:
        raise HTTPException(status_code=404, detail="Checklist item not found")
    return db.query(AuditItemAttachment).filter(AuditItemAttachment.item_id == item_id).order_by(AuditItemAttachment.uploaded_at.desc()).all()


@router.get("/checklist/attachments/{attachment_id}/download")
async def download_item_attachment(attachment_id: int, db: Session = Depends(get_db), current_user: User = Depends(get_current_user)):
    att = db.query(AuditItemAttachment).get(attachment_id)
    if not att:
        raise HTTPException(status_code=404, detail="Attachment not found")
    
    # Use storage service for secure file download
    storage_service = StorageService()
    return storage_service.create_file_response(att.file_path, att.filename)


@router.delete("/checklist/attachments/{attachment_id}", dependencies=[Depends(require_permission_dependency("audits:delete"))])
async def delete_item_attachment(attachment_id: int, db: Session = Depends(get_db), current_user: User = Depends(get_current_user)):
    att = db.query(AuditItemAttachment).get(attachment_id)
    if not att:
        raise HTTPException(status_code=404, detail="Attachment not found")
    
    # Use storage service to delete the file
    storage_service = StorageService()
    storage_service.delete_file(att.file_path)
    
    db.delete(att)
    db.commit()
    return {"message": "Attachment deleted"}


# List/download/delete attachments for finding
@router.get("/findings/{finding_id}/attachments", response_model=List[AuditFindingAttachmentResponse])
async def list_finding_attachments(finding_id: int, db: Session = Depends(get_db), current_user: User = Depends(get_current_user)):
    f = db.query(AuditFinding).get(finding_id)
    if not f:
        raise HTTPException(status_code=404, detail="Finding not found")
    return db.query(AuditFindingAttachment).filter(AuditFindingAttachment.finding_id == finding_id).order_by(AuditFindingAttachment.uploaded_at.desc()).all()


@router.get("/findings/attachments/{attachment_id}/download")
async def download_finding_attachment(attachment_id: int, db: Session = Depends(get_db), current_user: User = Depends(get_current_user)):
    att = db.query(AuditFindingAttachment).get(attachment_id)
    if not att:
        raise HTTPException(status_code=404, detail="Attachment not found")
    
    # Use storage service for secure file download
    storage_service = StorageService()
    return storage_service.create_file_response(att.file_path, att.filename)


@router.delete("/findings/attachments/{attachment_id}", dependencies=[Depends(require_permission_dependency("audits:delete"))])
async def delete_finding_attachment(attachment_id: int, db: Session = Depends(get_db), current_user: User = Depends(get_current_user)):
    att = db.query(AuditFindingAttachment).get(attachment_id)
    if not att:
        raise HTTPException(status_code=404, detail="Attachment not found")
    
    # Use storage service to delete the file
    storage_service = StorageService()
    storage_service.delete_file(att.file_path)
    
    db.delete(att)
    db.commit()
    return {"message": "Attachment deleted"}


# Create Non-Conformance from a finding
@router.post("/findings/{finding_id}/create-nc", response_model=NonConformanceResponse)
async def create_nc_from_finding(
    finding_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    finding = db.query(AuditFinding).get(finding_id)
    if not finding:
        raise HTTPException(status_code=404, detail="Finding not found")

    # Build NC data from finding
    nc_payload = NonConformanceCreate(
        title=f"Audit Finding NC - {finding.clause_ref or 'No Clause'}",
        description=finding.description,
        source=NonConformanceSource.AUDIT,
        severity=finding.severity.value if hasattr(finding.severity, 'value') else str(finding.severity),
        impact_area="food safety",
        category="audit",
        target_resolution_date=finding.target_completion_date,
    )

    # Create NC
    nc = NCModel(
        nc_number=f"NC-{int(datetime.utcnow().timestamp())}",
        title=nc_payload.title,
        description=nc_payload.description,
        source=NonConformanceSource.AUDIT,
        batch_reference=None,
        product_reference=None,
        process_reference=None,
        location=None,
        severity=nc_payload.severity,
        impact_area=nc_payload.impact_area,
        category=nc_payload.category,
        status=NonConformanceStatus.OPEN,
        reported_date=datetime.utcnow(),
        target_resolution_date=nc_payload.target_resolution_date,
        reported_by=current_user.id,
    )
    db.add(nc)
    db.commit()
    db.refresh(nc)

    # Link NC back to finding
    finding.related_nc_id = nc.id
    db.commit()

    try:
        audit_event(db, current_user.id, "nc_created_from_finding", "audits", str(finding.audit_id), {"nc_id": nc.id, "finding_id": finding_id})
    except Exception:
        pass

    return nc


# Stats endpoint
@router.get("/stats", response_model=AuditStatsResponse, dependencies=[Depends(require_permission_dependency("audits:view"))])
async def get_audit_stats(db: Session = Depends(get_db), current_user: User = Depends(get_current_user)):
    q = db.query(AuditModel)
    total = q.count()
    by_status = {s.value: db.query(AuditModel).filter(AuditModel.status == s).count() for s in AuditStatus}
    by_type = {t.value: db.query(AuditModel).filter(AuditModel.audit_type == t).count() for t in AuditType}
    recent = (
        db.query(AuditModel.id, AuditModel.title, AuditModel.created_at)
        .order_by(AuditModel.created_at.desc())
        .limit(10)
        .all()
    )
    recent_created = [
        {"id": str(r.id), "title": r.title, "created_at": r.created_at.isoformat() if r.created_at else ""}
        for r in recent
    ]
    return AuditStatsResponse(total=total, by_status=by_status, by_type=by_type, recent_created=recent_created)


@router.get("/kpis/overview", response_model=AuditKpisResponse, dependencies=[Depends(require_permission_dependency("audits:view"))])
async def get_audit_kpis_overview(
	period: str = Query("month", pattern="^(week|month|quarter|year)$"),
	department: Optional[str] = Query(None),
	auditor_id: Optional[int] = Query(None),
	db: Session = Depends(get_db),
	current_user: User = Depends(get_current_user)
):
	"""
	Get audit KPIs with filtering by period, department, and auditor.
	
	Args:
		period: Time period for KPI calculations (week|month|quarter|year)
		department: Filter by auditee department
		auditor_id: Filter by auditor ID (lead or team auditor)
	"""
	from datetime import datetime, timedelta
	
	# Calculate date range based on period
	now = datetime.utcnow()
	if period == "week":
		start_date = now - timedelta(days=7)
	elif period == "month":
		start_date = now - timedelta(days=30)
	elif period == "quarter":
		start_date = now - timedelta(days=90)
	elif period == "year":
		start_date = now - timedelta(days=365)
	else:
		start_date = now - timedelta(days=30)  # Default to month
	
	# Build base query with period filter
	q_audits = db.query(AuditModel).filter(AuditModel.created_at >= start_date)
	
	# Apply additional filters
	if department:
		q_audits = q_audits.filter(AuditModel.auditee_department == department)
	if auditor_id:
		q_audits = q_audits.filter((AuditModel.auditor_id == auditor_id) | (AuditModel.lead_auditor_id == auditor_id))

	# Get audits for the period
	audits = q_audits.all()
	
	# Lead time average: prefer plan.approved_at when available, fallback to audit.start_date
	lead_deltas = []
	for a in audits:
		plan = db.query(AuditPlan).filter(AuditPlan.audit_id == a.id).first()
		start_ref = plan.approved_at if plan and plan.approved_at else getattr(a, 'start_date', None)
		created_ref = getattr(a, 'created_at', None)
		if start_ref and created_ref:
			try:
				lead_deltas.append((start_ref - created_ref).days)
			except Exception:
				pass
	lead_time_days_avg = (sum(lead_deltas) / len(lead_deltas)) if lead_deltas else None

	# On-time rate: actual_end_at vs planned end_date
	completed = [a for a in audits if getattr(a, 'status', None) in {AuditStatus.COMPLETED, AuditStatus.CLOSED}]
	on_time_count = 0
	total_completed = 0
	for a in completed:
		planned_end = getattr(a, 'end_date', None)
		actual_end = getattr(a, 'actual_end_at', None) or getattr(a, 'updated_at', None)
		if planned_end and actual_end:
			total_completed += 1
			if actual_end <= planned_end:
				on_time_count += 1
	on_time_audit_rate = (on_time_count / total_completed) if total_completed else None

	# Finding closure days average - filter by period and auditor
	q_findings = db.query(AuditFinding).filter(AuditFinding.created_at >= start_date)
	if auditor_id:
		q_findings = q_findings.filter(AuditFinding.responsible_person_id == auditor_id)
	findings = q_findings.all()
	closure_deltas = []
	for f in findings:
		if f.closed_at and f.created_at:
			try:
				closure_deltas.append((f.closed_at - f.created_at).days)
			except Exception:
				pass
	finding_closure_days_avg = (sum(closure_deltas) / len(closure_deltas)) if closure_deltas else None

	# Additional KPI metrics
	total_audits = len(audits)
	completed_audits = len([a for a in audits if a.status in {AuditStatus.COMPLETED, AuditStatus.CLOSED}])
	
	# Overdue audits (past end_date but not completed)
	overdue_audits = 0
	for a in audits:
		if a.status not in {AuditStatus.COMPLETED, AuditStatus.CLOSED} and a.end_date and a.end_date < now:
			overdue_audits += 1
	
	# Finding metrics
	total_findings = len(findings)
	open_findings = len([f for f in findings if f.status in {FindingStatus.OPEN, FindingStatus.IN_PROGRESS}])
	
	# Overdue findings (past target_completion_date but not closed)
	overdue_findings = 0
	for f in findings:
		if f.status not in {FindingStatus.VERIFIED, FindingStatus.CLOSED} and f.target_completion_date and f.target_completion_date < now:
			overdue_findings += 1
	
	# Critical findings
	critical_findings = len([f for f in findings if f.severity == FindingSeverity.CRITICAL])

	return AuditKpisResponse(
		lead_time_days_avg=lead_time_days_avg,
		on_time_audit_rate=on_time_audit_rate,
		finding_closure_days_avg=finding_closure_days_avg,
		total_audits=total_audits,
		completed_audits=completed_audits,
		overdue_audits=overdue_audits,
		total_findings=total_findings,
		open_findings=open_findings,
		overdue_findings=overdue_findings,
		critical_findings=critical_findings,
		period=period,
		department=department,
		auditor_id=auditor_id,
	)


# Export list to PDF or XLSX
@router.post("/export", dependencies=[Depends(require_permission_dependency("audits:export"))])
async def export_audits(
    format: str = Query("pdf", pattern="^(pdf|xlsx)$"),
    search: Optional[str] = Query(None),
    audit_type: Optional[str] = Query(None),
    status: Optional[str] = Query(None),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    q = db.query(AuditModel)
    if audit_type:
        q = q.filter(AuditModel.audit_type == AuditType(audit_type))
    if status:
        q = q.filter(AuditModel.status == AuditStatus(status))
    if search:
        like = f"%{search}%"
        q = q.filter(AuditModel.title.ilike(like))
    items = q.order_by(AuditModel.created_at.desc()).all()

    if format == "xlsx":
        wb = Workbook(); ws = wb.active; ws.title = "Audits"
        ws.append(["ID", "Title", "Type", "Status", "Start", "End", "Lead Auditor", "Auditee Dept"]) 
        for a in items:
            ws.append([a.id, a.title, a.audit_type.value, a.status.value, str(a.start_date or ''), str(a.end_date or ''), a.lead_auditor_id or '', a.auditee_department or ''])
        buf = BytesIO(); wb.save(buf); buf.seek(0)
        headers = {"Content-Disposition": "attachment; filename=audits.xlsx"}
        return StreamingResponse(buf, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers=headers)
    else:
        buf = BytesIO(); c = canvas.Canvas(buf, pagesize=A4)
        x, y = 40, 800
        c.setFont("Helvetica-Bold", 14); c.drawString(x, y, "Audit List"); y -= 20; c.setFont("Helvetica", 10)
        for a in items:
            line = f"#{a.id} {a.title} | {a.audit_type.value} | {a.status.value}"
            c.drawString(x, y, line)
            y -= 14
            if y < 60: c.showPage(); y = 800
        c.showPage(); c.save(); buf.seek(0)
        headers = {"Content-Disposition": "attachment; filename=audits.pdf"}
        return StreamingResponse(buf, media_type="application/pdf", headers=headers)


@router.post("/reports/consolidated", dependencies=[Depends(require_permission_dependency("audits:export"))])
async def export_consolidated_reports(
    format: str = Query("pdf", pattern="^(pdf|xlsx)$"),
    date_from: Optional[datetime] = Query(None),
    date_to: Optional[datetime] = Query(None),
    program_id: Optional[int] = Query(None),
    department: Optional[str] = Query(None),
    auditor_id: Optional[int] = Query(None),
    status: Optional[str] = Query(None),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    q = db.query(AuditModel)
    if date_from:
        q = q.filter(AuditModel.created_at >= date_from)
    if date_to:
        q = q.filter(AuditModel.created_at <= date_to)
    if program_id:
        q = q.filter(AuditModel.program_id == program_id)
    if department:
        q = q.filter(AuditModel.auditee_department == department)
    if auditor_id:
        q = q.filter((AuditModel.auditor_id == auditor_id) | (AuditModel.lead_auditor_id == auditor_id))
    if status:
        q = q.filter(AuditModel.status == AuditStatus(status))
    items = q.order_by(AuditModel.created_at.desc()).all()

    if format == "xlsx":
        wb = Workbook(); ws = wb.active; ws.title = "Consolidated Reports"
        ws.append(["ID", "Title", "Type", "Status", "Start", "End", "Department", "Lead Auditor", "Auditor"]) 
        for a in items:
            ws.append([a.id, a.title, a.audit_type.value, a.status.value, str(a.start_date or ''), str(a.end_date or ''), a.auditee_department or '', a.lead_auditor_id or '', a.auditor_id or ''])
        buf = BytesIO(); wb.save(buf); buf.seek(0)
        headers = {"Content-Disposition": "attachment; filename=audits_consolidated.xlsx"}
        return StreamingResponse(buf, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers=headers)
    else:
        buf = BytesIO(); c = canvas.Canvas(buf, pagesize=A4)
        x, y = 40, 800
        c.setFont("Helvetica-Bold", 14); c.drawString(x, y, "Consolidated Audit Reports"); y -= 20; c.setFont("Helvetica", 10)
        for a in items:
            line = f"#{a.id} {a.title} | {a.audit_type.value} | {a.status.value} | Dept={a.auditee_department or ''}"
            c.drawString(x, y, line)
            y -= 14
            if y < 60: c.showPage(); y = 800
        c.showPage(); c.save(); buf.seek(0)
        headers = {"Content-Disposition": "attachment; filename=audits_consolidated.pdf"}
        return StreamingResponse(buf, media_type="application/pdf", headers=headers)


# Single audit report export
@router.get("/{audit_id:int}/report", dependencies=[Depends(require_permission_dependency("audits:export"))])
async def export_audit_report(
    audit_id: int,
    format: str = Query("pdf", pattern="^(pdf|xlsx)$"),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    audit = db.query(AuditModel).get(audit_id)
    if not audit:
        raise HTTPException(status_code=404, detail="Audit not found")
    items = db.query(AuditChecklistItem).filter(AuditChecklistItem.audit_id == audit_id).all()
    findings = db.query(AuditFinding).filter(AuditFinding.audit_id == audit_id).all()

    if format == "xlsx":
        wb = Workbook(); ws = wb.active; ws.title = "Audit Report"
        ws.append(["Audit", audit.title])
        ws.append(["Type", audit.audit_type.value]); ws.append(["Status", audit.status.value])
        ws.append([]); ws.append(["Checklist Items"])
        ws.append(["Clause", "Question", "Response", "Score", "Comment", "Responsible", "Due Date"])
        for i in items:
            ws.append([i.clause_ref or '', (i.question or '')[:100], i.response.value if i.response else '', i.score or '', i.comment or '', i.responsible_person_id or '', str(i.due_date or '')])
        ws.append([]); ws.append(["Findings"])
        ws.append(["Clause", "Description", "Severity", "Status", "Target Date", "Related NC"])
        for f in findings:
            ws.append([f.clause_ref or '', (f.description or '')[:120], f.severity.value, f.status.value, str(f.target_completion_date or ''), f.related_nc_id or ''])
        buf = BytesIO(); wb.save(buf); buf.seek(0)
        headers = {"Content-Disposition": f"attachment; filename=audit_{audit_id}.xlsx"}
        return StreamingResponse(buf, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers=headers)
    else:
        buf = BytesIO(); c = canvas.Canvas(buf, pagesize=A4)
        x, y = 40, 800
        c.setFont("Helvetica-Bold", 14); c.drawString(x, y, f"Audit Report: {audit.title}"); y -= 18
        c.setFont("Helvetica", 10)
        c.drawString(x, y, f"Type: {audit.audit_type.value} | Status: {audit.status.value}"); y -= 16
        c.drawString(x, y, "Checklist Items:"); y -= 14
        for i in items:
            line = f"[{i.clause_ref or '-'}] {(i.question or '')[:70]} | {i.response.value if i.response else ''} | score={i.score or ''}"
            c.drawString(x, y, line); y -= 12
            if y < 60: c.showPage(); y = 800
        if y < 100: c.showPage(); y = 800
        c.drawString(40, y, "Findings:"); y -= 14
        for f in findings:
            line = f"[{f.clause_ref or '-'}] {(f.description or '')[:70]} | sev={f.severity.value} | status={f.status.value} | NC={f.related_nc_id or ''}"
            c.drawString(40, y, line); y -= 12
            if y < 60: c.showPage(); y = 800
        c.showPage(); c.save(); buf.seek(0)
        headers = {"Content-Disposition": f"attachment; filename=audit_{audit_id}.pdf"}
        return StreamingResponse(buf, media_type="application/pdf", headers=headers)


# Report approval and history
@router.post("/{audit_id:int}/report/approve", dependencies=[Depends(require_permission_dependency("audits:approve"))])
async def approve_audit_report(
    audit_id: int,
    payload: AuditReportApproveRequest,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    audit = db.query(AuditModel).get(audit_id)
    if not audit:
        raise HTTPException(status_code=404, detail="Audit not found")

    # Ownership check for approval
    check_audit_ownership(audit, current_user, db, "update")

    # Determine next version
    last = (
        db.query(AuditReportHistory)
        .filter(AuditReportHistory.audit_id == audit_id)
        .order_by(AuditReportHistory.version.desc())
        .first()
    )
    next_version = 1 if not last else last.version + 1

    rec = AuditReportHistory(
        audit_id=audit_id,
        version=next_version,
        approved_by=current_user.id,
        approved_at=datetime.utcnow(),
        notes=payload.notes,
        file_path=payload.file_path,
    )
    db.add(rec)
    db.commit()
    db.refresh(rec)
    try:
        audit_event(db, current_user.id, "audit_report_approved", "audits", str(audit_id), {"version": next_version})
    except Exception:
        pass
    return {"audit_id": audit_id, "version": next_version, "approved_by": current_user.id, "approved_at": rec.approved_at}


@router.get("/{audit_id:int}/report/history", response_model=AuditReportHistoryResponse, dependencies=[Depends(require_permission_dependency("audits:view"))])
async def get_audit_report_history(
    audit_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    audit = db.query(AuditModel).get(audit_id)
    if not audit:
        raise HTTPException(status_code=404, detail="Audit not found")

    # Basic view access
    check_audit_ownership(audit, current_user, db, "view")
    items = (
        db.query(AuditReportHistory)
        .filter(AuditReportHistory.audit_id == audit_id)
        .order_by(AuditReportHistory.version.desc())
        .all()
    )
    return {"items": items}
# Auditees
@router.get("/{audit_id:int}/auditees", response_model=List[AuditeeResponse], dependencies=[Depends(require_permission_dependency("audits:view"))])
async def list_auditees(audit_id: int, db: Session = Depends(get_db), current_user: User = Depends(get_current_user)):
    audit = db.query(AuditModel).get(audit_id)
    if not audit:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Audit not found")
    
    # Check ownership for view access
    check_audit_ownership(audit, current_user, db, "view")
    
    return db.query(AuditAuditee).filter(AuditAuditee.audit_id == audit_id).order_by(AuditAuditee.added_at.desc()).all()


@router.post("/{audit_id:int}/auditees", response_model=AuditeeResponse, dependencies=[Depends(require_permission_dependency("audits:update"))])
async def add_auditee(audit_id: int, user_id: int = Query(...), role: Optional[str] = Query(None), db: Session = Depends(get_db), current_user: User = Depends(get_current_user)):
    audit = db.query(AuditModel).get(audit_id)
    if not audit:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Audit not found")
    
    # Check ownership for update access
    check_audit_ownership(audit, current_user, db, "update")
    
    aa = AuditAuditee(audit_id=audit_id, user_id=user_id, role=role)
    db.add(aa)
    db.commit()
    db.refresh(aa)
    try:
        audit_event(db, current_user.id, "audit_auditee_added", "audits", str(audit_id), {"user_id": user_id})
    except Exception:
        pass
    return aa


@router.delete("/auditees/{id}", dependencies=[Depends(require_permission_dependency("audits:delete"))])
async def remove_auditee(id: int, db: Session = Depends(get_db), current_user: User = Depends(get_current_user)):
    aa = db.query(AuditAuditee).get(id)
    if not aa:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Auditee not found")
    
    # Get the audit to check ownership
    audit = db.query(AuditModel).get(aa.audit_id)
    if not audit:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Audit not found")
    
    # Check ownership for delete access
    check_audit_ownership(audit, current_user, db, "delete")
    
    audit_id = aa.audit_id
    db.delete(aa)
    db.commit()
    try:
        audit_event(db, current_user.id, "audit_auditee_removed", "audits", str(audit_id), {"auditee_id": id})
    except Exception:
        pass
    return {"message": "Auditee removed"}


# Audit Plan endpoints
@router.post("/{audit_id:int}/plan", response_model=AuditPlanResponse, dependencies=[Depends(require_permission_dependency("audits:update"))])
async def create_or_update_plan(
	audit_id: int,
	payload: AuditPlanCreate,
	db: Session = Depends(get_db),
	current_user: User = Depends(get_current_user)
):
	audit = db.query(AuditModel).get(audit_id)
	if not audit:
		raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Audit not found")
	
	# Check ownership for update access
	check_audit_ownership(audit, current_user, db, "update")
	
	plan = db.query(AuditPlan).filter(AuditPlan.audit_id == audit_id).first()
	if plan is None:
		plan = AuditPlan(audit_id=audit_id)
		db.add(plan)
	for field, value in payload.model_dump(exclude_unset=True).items():
		setattr(plan, field, value)
	db.commit()
	db.refresh(plan)
	return plan


@router.post("/{audit_id:int}/plan/approve", response_model=AuditPlanResponse, dependencies=[Depends(require_permission_dependency("audits:approve"))])
async def approve_plan(
	audit_id: int,
	db: Session = Depends(get_db),
	current_user: User = Depends(get_current_user)
):
	audit = db.query(AuditModel).get(audit_id)
	if not audit:
		raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Audit not found")
	
	# Check ownership for approve access (only lead auditor or program manager)
	check_audit_ownership(audit, current_user, db, "update")
	
	plan = db.query(AuditPlan).filter(AuditPlan.audit_id == audit_id).first()
	if not plan:
		raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Audit plan not found")
	plan.approved_by = current_user.id
	plan.approved_at = datetime.utcnow()
	db.commit()
	db.refresh(plan)
	return plan


@router.get("/{audit_id:int}/plan", response_model=AuditPlanResponse, dependencies=[Depends(require_permission_dependency("audits:view"))])
async def get_plan(
	audit_id: int,
	db: Session = Depends(get_db),
	current_user: User = Depends(get_current_user)
):
	audit = db.query(AuditModel).get(audit_id)
	if not audit:
		raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Audit not found")
	
	# Check ownership for view access
	check_audit_ownership(audit, current_user, db, "view")
	
	plan = db.query(AuditPlan).filter(AuditPlan.audit_id == audit_id).first()
	if not plan:
		raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Audit plan not found")
	return plan


# Audit Program CRUD endpoints
@router.get("/programs", response_model=AuditProgramListResponse, dependencies=[Depends(require_permission_dependency("audits:view"))])
async def list_audit_programs(
    skip: int = Query(0, ge=0),
    limit: int = Query(100, ge=1, le=1000),
    status: Optional[str] = Query(None, description="Filter by program status"),
    year: Optional[int] = Query(None, description="Filter by year"),
    manager_id: Optional[int] = Query(None, description="Filter by program manager"),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """List audit programs with pagination and filtering"""
    query = db.query(AuditProgram)
    
    # Apply filters
    if status:
        query = query.filter(AuditProgram.status == status)
    if year:
        query = query.filter(AuditProgram.year == year)
    if manager_id:
        query = query.filter(AuditProgram.manager_id == manager_id)
    
    # Get total count
    total = query.count()
    
    # Apply pagination
    programs = query.offset(skip).limit(limit).all()
    
    # Calculate pagination info
    pages = (total + limit - 1) // limit
    page = (skip // limit) + 1
    
    return AuditProgramListResponse(
        items=programs,
        total=total,
        page=page,
        size=limit,
        pages=pages
    )


@router.post("/programs", response_model=AuditProgramResponse, dependencies=[Depends(require_permission_dependency("audits:create"))])
async def create_audit_program(
    payload: AuditProgramCreate,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """Create a new audit program"""
    program = AuditProgram(
        **payload.model_dump(),
        created_by=current_user.id
    )
    db.add(program)
    db.commit()
    db.refresh(program)
    
    try:
        audit_event(db, current_user.id, "audit_program_created", "audit_programs", str(program.id), payload.model_dump())
    except Exception:
        pass
    
    return program


@router.get("/programs/{program_id}", response_model=AuditProgramResponse, dependencies=[Depends(require_permission_dependency("audits:view"))])
async def get_audit_program(
    program_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """Get a specific audit program by ID"""
    program = db.query(AuditProgram).get(program_id)
    if not program:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Audit program not found")
    
    return program


@router.put("/programs/{program_id}", response_model=AuditProgramResponse, dependencies=[Depends(require_permission_dependency("audits:update"))])
async def update_audit_program(
    program_id: int,
    payload: AuditProgramUpdate,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """Update an audit program"""
    program = db.query(AuditProgram).get(program_id)
    if not program:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Audit program not found")
    
    # Check if user is program manager or has manage_program permission
    if program.manager_id != current_user.id and not check_permission(current_user.id, Module.AUDITS, PermissionType.MANAGE_PROGRAM, db):
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="Only the program manager or audit program manager can update this program"
        )
    
    # Update fields
    for field, value in payload.model_dump(exclude_unset=True).items():
        setattr(program, field, value)
    
    program.updated_at = datetime.utcnow()
    db.commit()
    db.refresh(program)
    
    try:
        audit_event(db, current_user.id, "audit_program_updated", "audit_programs", str(program_id), payload.model_dump(exclude_unset=True))
    except Exception:
        pass
    
    return program


@router.delete("/programs/{program_id}", dependencies=[Depends(require_permission_dependency("audits:delete"))])
async def delete_audit_program(
    program_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """Delete an audit program"""
    program = db.query(AuditProgram).get(program_id)
    if not program:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Audit program not found")
    
    # Check if user is program manager or has manage_program permission
    if program.manager_id != current_user.id and not check_permission(current_user.id, Module.AUDITS, PermissionType.MANAGE_PROGRAM, db):
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="Only the program manager or audit program manager can delete this program"
        )
    
    # Check if program has associated audits
    audit_count = db.query(AuditModel).filter(AuditModel.program_id == program_id).count()
    if audit_count > 0:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"Cannot delete program with {audit_count} associated audits. Remove or reassign audits first."
        )
    
    db.delete(program)
    db.commit()
    
    try:
        audit_event(db, current_user.id, "audit_program_deleted", "audit_programs", str(program_id), {"program_name": program.name})
    except Exception:
        pass
    
    return {"message": "Audit program deleted successfully"}


@router.get("/programs/{program_id}/schedule", dependencies=[Depends(require_permission_dependency("audits:view"))])
async def get_program_schedule(
    program_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """Get the schedule for a specific audit program"""
    program = db.query(AuditProgram).get(program_id)
    if not program:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Audit program not found")
    
    # Get all audits in this program
    audits = db.query(AuditModel).filter(AuditModel.program_id == program_id).all()
    
    # Get program schedule if available
    schedule = program.schedule or {}
    
    return {
        "program_id": program_id,
        "program_name": program.name,
        "schedule": schedule,
        "audits": [
            {
                "id": audit.id,
                "title": audit.title,
                "start_date": audit.start_date,
                "end_date": audit.end_date,
                "status": audit.status,
                "auditor_id": audit.auditor_id,
                "lead_auditor_id": audit.lead_auditor_id
            }
            for audit in audits
        ],
        "total_audits": len(audits),
        "completed_audits": len([a for a in audits if a.status == AuditStatus.COMPLETED]),
        "planned_audits": len([a for a in audits if a.status == AuditStatus.PLANNED]),
        "in_progress_audits": len([a for a in audits if a.status == AuditStatus.IN_PROGRESS])
    }


@router.get("/programs/{program_id}/kpis", response_model=AuditProgramKpisResponse, dependencies=[Depends(require_permission_dependency("audits:view"))])
async def get_program_kpis(
    program_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """Get KPIs for a specific audit program"""
    program = db.query(AuditProgram).get(program_id)
    if not program:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Audit program not found")
    
    # Get all audits in this program
    audits = db.query(AuditModel).filter(AuditModel.program_id == program_id).all()
    
    # Calculate KPIs
    total_audits = len(audits)
    completed_audits = len([a for a in audits if a.status == AuditStatus.COMPLETED])
    overdue_audits = len([a for a in audits if a.end_date and a.end_date < datetime.utcnow() and a.status != AuditStatus.COMPLETED])
    
    # Calculate on-time rate
    on_time_rate = None
    if total_audits > 0:
        on_time_rate = completed_audits / total_audits
    
    # Get all findings from audits in this program
    audit_ids = [a.id for a in audits]
    findings = db.query(AuditFinding).filter(AuditFinding.audit_id.in_(audit_ids)).all()
    
    total_findings = len(findings)
    open_findings = len([f for f in findings if f.status != FindingStatus.CLOSED])
    critical_findings = len([f for f in findings if f.severity == FindingSeverity.CRITICAL])
    
    # Calculate average closure days
    closed_findings = [f for f in findings if f.closed_at and f.created_at]
    average_closure_days = None
    if closed_findings:
        total_days = sum((f.closed_at - f.created_at).days for f in closed_findings)
        average_closure_days = total_days / len(closed_findings)
    
    return AuditProgramKpisResponse(
        program_id=program_id,
        program_name=program.name,
        total_audits=total_audits,
        completed_audits=completed_audits,
        overdue_audits=overdue_audits,
        on_time_rate=on_time_rate,
        total_findings=total_findings,
        open_findings=open_findings,
        critical_findings=critical_findings,
        average_closure_days=average_closure_days
    )


# Risk-Based Planning endpoints
@router.get("/programs/{program_id}/risks", response_model=AuditRiskListResponse, dependencies=[Depends(require_permission_dependency("audits:view"))])
async def list_program_risks(
    program_id: int,
    skip: int = Query(0, ge=0),
    limit: int = Query(100, ge=1, le=1000),
    risk_rating: Optional[str] = Query(None, description="Filter by risk rating"),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """List risks for a specific audit program"""
    # Verify program exists
    program = db.query(AuditProgram).get(program_id)
    if not program:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Audit program not found")
    
    query = db.query(AuditRisk).filter(AuditRisk.program_id == program_id)
    
    # Apply filters
    if risk_rating:
        query = query.filter(AuditRisk.risk_rating == risk_rating)
    
    # Get total count
    total = query.count()
    
    # Apply pagination
    risks = query.offset(skip).limit(limit).all()
    
    # Calculate pagination info
    pages = (total + limit - 1) // limit
    page = (skip // limit) + 1
    
    return AuditRiskListResponse(
        items=risks,
        total=total,
        page=page,
        size=limit,
        pages=pages
    )


@router.post("/programs/{program_id}/risks", response_model=AuditRiskResponse, dependencies=[Depends(require_permission_dependency("audits:create"))])
async def create_program_risk(
    program_id: int,
    payload: AuditRiskCreate,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """Create a new risk for an audit program"""
    # Verify program exists
    program = db.query(AuditProgram).get(program_id)
    if not program:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Audit program not found")
    
    risk = AuditRisk(
        **payload.model_dump(),
        program_id=program_id,
        created_by=current_user.id
    )
    db.add(risk)
    db.commit()
    db.refresh(risk)
    
    try:
        audit_event(db, current_user.id, "audit_risk_created", "audit_risks", str(risk.id), payload.model_dump())
    except Exception:
        pass
    
    return risk


@router.get("/programs/{program_id}/risks/{risk_id}", response_model=AuditRiskResponse, dependencies=[Depends(require_permission_dependency("audits:view"))])
async def get_program_risk(
    program_id: int,
    risk_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """Get a specific risk for an audit program"""
    risk = db.query(AuditRisk).filter(
        AuditRisk.id == risk_id,
        AuditRisk.program_id == program_id
    ).first()
    
    if not risk:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Risk not found")
    
    return risk


@router.put("/programs/{program_id}/risks/{risk_id}", response_model=AuditRiskResponse, dependencies=[Depends(require_permission_dependency("audits:update"))])
async def update_program_risk(
    program_id: int,
    risk_id: int,
    payload: AuditRiskUpdate,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """Update a risk for an audit program"""
    risk = db.query(AuditRisk).filter(
        AuditRisk.id == risk_id,
        AuditRisk.program_id == program_id
    ).first()
    
    if not risk:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Risk not found")
    
    # Update fields
    for field, value in payload.model_dump(exclude_unset=True).items():
        setattr(risk, field, value)
    
    risk.updated_at = datetime.utcnow()
    db.commit()
    db.refresh(risk)
    
    try:
        audit_event(db, current_user.id, "audit_risk_updated", "audit_risks", str(risk_id), payload.model_dump(exclude_unset=True))
    except Exception:
        pass
    
    return risk


@router.delete("/programs/{program_id}/risks/{risk_id}", dependencies=[Depends(require_permission_dependency("audits:delete"))])
async def delete_program_risk(
    program_id: int,
    risk_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """Delete a risk from an audit program"""
    risk = db.query(AuditRisk).filter(
        AuditRisk.id == risk_id,
        AuditRisk.program_id == program_id
    ).first()
    
    if not risk:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Risk not found")
    
    db.delete(risk)
    db.commit()
    
    try:
        audit_event(db, current_user.id, "audit_risk_deleted", "audit_risks", str(risk_id), {"area_name": risk.area_name})
    except Exception:
        pass
    
    return {"message": "Risk deleted successfully"}


@router.get("/programs/{program_id}/risk-plan", response_model=RiskPlanResponse, dependencies=[Depends(require_permission_dependency("audits:view"))])
async def get_program_risk_plan(
    program_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """Get risk-based audit plan for a program"""
    program = db.query(AuditProgram).get(program_id)
    if not program:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Audit program not found")
    
    # Get all risks for this program
    risks = db.query(AuditRisk).filter(AuditRisk.program_id == program_id).all()
    
    # Calculate risk statistics
    total_risks = len(risks)
    high_critical_risks = len([r for r in risks if r.risk_rating in [RiskLevel.HIGH, RiskLevel.CRITICAL]])
    overdue_audits = len([r for r in risks if r.next_audit_due and r.next_audit_due < datetime.utcnow()])
    
    # Generate suggested audits based on risk and elapsed time
    suggested_audits = []
    for risk in risks:
        # Calculate priority score based on risk rating and time since last audit
        priority_score = 0
        
        # Risk rating score
        risk_scores = {RiskLevel.LOW: 1, RiskLevel.MEDIUM: 2, RiskLevel.HIGH: 3, RiskLevel.CRITICAL: 4}
        priority_score += risk_scores.get(risk.risk_rating, 2)
        
        # Time factor (if overdue, increase priority)
        if risk.next_audit_due and risk.next_audit_due < datetime.utcnow():
            priority_score += 3  # High priority for overdue audits
        elif risk.last_audit_date:
            # Calculate months since last audit
            months_since = (datetime.utcnow() - risk.last_audit_date).days / 30
            if risk.audit_frequency_months and months_since > risk.audit_frequency_months:
                priority_score += 2  # Medium priority for past due audits
        
        suggested_audits.append({
            "risk_id": risk.id,
            "area_name": risk.area_name,
            "process_name": risk.process_name,
            "risk_rating": risk.risk_rating,
            "priority_score": priority_score,
            "last_audit_date": risk.last_audit_date,
            "next_audit_due": risk.next_audit_due,
            "suggested_audit_date": risk.next_audit_due or datetime.utcnow(),
            "responsible_auditor_id": risk.responsible_auditor_id,
            "rationale": risk.rationale
        })
    
    # Sort by priority score (highest first)
    suggested_audits.sort(key=lambda x: x["priority_score"], reverse=True)
    
    # Calculate risk coverage
    risk_coverage = {
        "total_areas": total_risks,
        "high_critical_coverage": high_critical_risks,
        "overdue_coverage": overdue_audits,
        "risk_distribution": {
            "low": len([r for r in risks if r.risk_rating == RiskLevel.LOW]),
            "medium": len([r for r in risks if r.risk_rating == RiskLevel.MEDIUM]),
            "high": len([r for r in risks if r.risk_rating == RiskLevel.HIGH]),
            "critical": len([r for r in risks if r.risk_rating == RiskLevel.CRITICAL])
        }
    }
    
    return RiskPlanResponse(
        program_id=program_id,
        program_name=program.name,
        risks=risks,
        suggested_audits=suggested_audits,
        risk_coverage=risk_coverage,
        total_risks=total_risks,
        high_critical_risks=high_critical_risks,
        overdue_audits=overdue_audits
    )


def assess_competence(user_id: int, required_competencies: List[str], db: Session) -> CompetenceAssessmentResponse:
    """
    Assess user competence for audit assignment
    This is a simplified implementation - in production, this would integrate with the Training module
    """
    # TODO: Integrate with actual Training module to check completed trainings
    # For now, we'll simulate competence assessment
    
    # Simulate user having some competencies
    user_competencies = ["HACCP", "ISO 22000", "Food Safety"]  # This would come from Training module
    
    missing_competencies = [comp for comp in required_competencies if comp not in user_competencies]
    training_recommendations = missing_competencies.copy()
    
    # Calculate competence score (0-100)
    competence_score = max(0, 100 - (len(missing_competencies) * 25))
    
    # Determine if user can be assigned
    can_assign = len(missing_competencies) <= 1  # Allow assignment if missing 1 or fewer competencies
    
    assessment_status = "competent" if can_assign else "needs_training"
    
    return CompetenceAssessmentResponse(
        user_id=user_id,
        assessment_status=assessment_status,
        missing_competencies=missing_competencies,
        training_recommendations=training_recommendations,
        competence_score=competence_score,
        can_assign=can_assign,
        notes=f"User has {len(user_competencies)} competencies, needs {len(missing_competencies)} more"
    )


def check_impartiality(auditor_id: int, auditee_department: str, db: Session) -> ImpartialityCheckResponse:
    """
    Check auditor impartiality for audit assignment
    """
    # Get auditor's department from user record
    auditor = db.query(User).get(auditor_id)
    if not auditor:
        return ImpartialityCheckResponse(
            auditor_id=auditor_id,
            is_impartial=False,
            conflict_type="auditor_not_found",
            requires_approval=True,
            approval_level="management",
            notes="Auditor not found in system"
        )
    
    auditor_department = getattr(auditor, 'department', None)
    
    # Check for department conflict
    if auditor_department and auditor_department.lower() == auditee_department.lower():
        return ImpartialityCheckResponse(
            auditor_id=auditor_id,
            is_impartial=False,
            conflict_type="department_conflict",
            requires_approval=True,
            approval_level="management",
            notes=f"Auditor and auditee are in same department: {auditor_department}"
        )
    
    # Check for recent involvement (simplified - would check audit history)
    # TODO: Check if auditor was recently involved with auditee department
    
    return ImpartialityCheckResponse(
        auditor_id=auditor_id,
        is_impartial=True,
        conflict_type=None,
        requires_approval=False,
        approval_level=None,
        notes="No conflicts detected"
    )


